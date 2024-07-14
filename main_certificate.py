import os
import re
import csv
from certificate import *
from docx import Document
from docx2pdf import convert
from openpyxl import load_workbook

mailerpath = "Data/Mail.xlsm"
htmltemplatepath = "Data/mailtemplate.html"

# Create output folders if they don't exist
os.makedirs("Output/Doc", exist_ok=True)
os.makedirs("Output/PDF", exist_ok=True)

# Function to update the mailer Excel sheet
def updatemailer(row, workbook, sheet, email, filepath, sub, body, status, cc=""):
    sheet.cell(row=row, column=1).value = email
    sheet.cell(row=row, column=2).value = cc
    sheet.cell(row=row, column=3).value = sub
    sheet.cell(row=row, column=4).value = body
    sheet.cell(row=row, column=5).value = filepath
    sheet.cell(row=row, column=6).value = status
    workbook.save(filename=mailerpath)

# Function to get the workbook and sheet
def getworkbook(filename):
    wb = load_workbook(filename=filename, read_only=False, keep_vba=True)
    sheet = wb.active
    return wb, sheet

# Function to get the HTML template
def gethtmltemplate(htmltemplatepath=htmltemplatepath):
    with open(htmltemplatepath, "r") as file:
        return file.read()

# Function to create the email subject and body
def getmail(name, event, ambassador):
    sub = f"Certificate of Participation for {name}"
    html = gethtmltemplate(htmltemplatepath)
    body = html.format(name=name, event=event, ambassador=ambassador)
    return sub, body

# Function to get participants from a CSV file
def get_participants(f):
    data = []
    with open(f, mode="r", encoding='iso-8859-1') as file:
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            data.append(row)
    return data

# Function to create the certificates and update the mailer Excel sheet
def create_docx_files(filename, list_participate):

    wb, sheet = getworkbook(mailerpath)

    event = input("Enter the event name: ")
    ambassador = input("Enter Ambassador Name: ")

    for index, participate in enumerate(list_participate):
        # Use original file every time
        doc = Document(filename)
        name_key = next(key for key in participate if re.search(r'\bName\b', key, re.IGNORECASE))
        email_key = next(key for key in participate if re.search(r'\bEmail\b', key, re.IGNORECASE))
        name = participate[name_key]
        email = participate[email_key]

        replace_participant_name(doc, name)
        replace_event_name(doc, event)
        replace_ambassador_name(doc, ambassador)

        doc_path = f'Output/Doc/{name}.docx'
        pdf_path = f'Output/PDF/{name}.pdf'
        
        doc.save(doc_path)
        print(f"Creating {pdf_path}")
        convert(doc_path, pdf_path)

        sub, body = getmail(name, event, ambassador)
        filepath = os.path.abspath(pdf_path)

        updatemailer(row=index+2, workbook=wb, sheet=sheet, email=email, filepath=filepath, sub=sub, body=body, status="Send")

# Main script execution
if __name__ == "__main__":
    # Get certificate template path
    certificate_file = "Data/Event Certificate Template.docx"
    
    # Get participants path
    participate_file = "Data/"+("ParticipantList.csv" if (input("Test Mode (Y/N): ").lower())[0] == "n" else "temp.csv")

    # Get participants
    list_participate = get_participants(participate_file)

    # Process data
    create_docx_files(certificate_file, list_participate)
